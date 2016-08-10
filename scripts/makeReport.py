#!/usr/bin/python
# coding: utf-8 

"""Script creant un rapport complet au format xls puis en pdf.
Ludovic KOSTHOWA (Debut : 06/04/16)
Info: Creation en cours, script peut etre modifie a tout moment.
"""

from openpyxl import *
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.cell import get_column_letter
import os
import codecs

class MakeReport:
	def __init__(self,REPERTORYVCF,i,RESULTDIR,pathREPERTORYVCF):
		self.report_body(REPERTORYVCF,i,RESULTDIR,pathREPERTORYVCF)



	def report_body(self,REPERTORYVCF,i,RESULTDIR,pathREPERTORYVCF):
		avecHotspot=False
		if i[-3:]=="vcf":
			i=i[:-4]
			#dans le cas d'un titre avec NOMPATIENT_v1
			title = i[:-5] 
		else: title =i
		mycwd = os.getcwd()

		report = open(RESULTDIR+"/"+REPERTORYVCF+"/temp/Report_info_"+i+".txt", "w")
		report.write("\n")
		report.write("Rapport de "+title+"\n\n")
		report.write("\n")
		infos = MakeReport.informations(self,REPERTORYVCF,i,RESULTDIR,pathREPERTORYVCF)
		if infos != None:
			report.write(infos)
			report.write("\n")
		if os.path.exists(RESULTDIR+"/"+REPERTORYVCF+"/"+REPERTORYVCF+"_globalInformations.txt") == True:
			report.write("Informations:")
			report.write("\n")
			report.write("\n")
			File = RESULTDIR+"/"+REPERTORYVCF+"/"+REPERTORYVCF+"_globalInformations.txt"
			with open(File,'r') as file:
				file = file.readlines()
				legendes = str(file[0])
				legendesReplace = legendes.replace(", ","\t")
				legendesReplace = legendesReplace.replace("\n","")
				report.write(legendesReplace)
				report.write("\n")
				for element in file:
					element = element.replace(", ","\t")
					element1 = element.split("\t")
					if element1[0] in i:
						report.write(element)
					elif element1[1] in i:
						report.write(element)
			report.write("\n")

		report = open(RESULTDIR+"/"+REPERTORYVCF+"/temp/Report_"+i+".txt", "w")
		notAlreadyDone=True
		if os.path.exists(RESULTDIR+"/"+REPERTORYVCF+"/temp/HSm_"+i+".vcf") == True:
			notAlreadyDone=False
			report.write("Dans les Hotspots:")
			report.write("\n")
			report.write("Variants:")
			report.write("\n")
			File = RESULTDIR+"/"+REPERTORYVCF+"/temp/HSm_"+i+".vcf"
			with open(File,'r') as file:
				file = file.readlines()
				for element in file:
					report.write(element)
			report.write("\n")
		if os.path.exists(RESULTDIR+"/"+REPERTORYVCF+"/temp/HSm_questionable_"+i+".vcf") == True:
			if notAlreadyDone:
				report.write("Dans les Hotspots")
				report.write("\n")
			report.write("Variants détectés mais non retenus: (NOCALL ou allele_freq < 1%  ou < 25 reads)")
			report.write("\n")
			File = RESULTDIR+"/"+REPERTORYVCF+"/temp/HSm_questionable_"+i+".vcf"
			with open(File,'r') as file:
				file = file.readlines()
				for element in file:
					report.write(element)
			report.write("\n")
		if os.path.exists(RESULTDIR+"/"+REPERTORYVCF+"/temp/nonMutatedHS_"+i+".vcf")== True:
			avecHotspot=True
			report.write("Profondeur Hotspots:")
			report.write("\n")
			File = RESULTDIR+"/"+REPERTORYVCF+"/temp/nonMutatedHS_"+i+".vcf"
			with open(File,'r') as file:
				file = file.readlines()
				for element in file:
					report.write(element)
			report.write("\n")
			report.write("\n")
		##########################################
		###TODO a virer
		###################
		if avecHotspot:
			report.write("Hors Hotspots:")
			report.write("\n")
		if os.path.exists(RESULTDIR+"/"+REPERTORYVCF+"/temp/mutations_"+i+".vcf")== True:
			report.write("Variants:")
			report.write("\n")
			File = RESULTDIR+"/"+REPERTORYVCF+"/temp/mutations_"+i+".vcf"
			with open(File,'r') as file:
				file = file.readlines()
				for element in file:
					report.write(element)
			report.write("\n")
		if os.path.exists(RESULTDIR+"/"+REPERTORYVCF+"/temp/uncertain_mutation_"+i+".vcf")== True:
			report.write("Variants détectés mais avec faible couverture: (cov < 300)")
			report.write("\n")
			File = RESULTDIR+"/"+REPERTORYVCF+"/temp/uncertain_mutation_"+i+".vcf"
			with open(File,'r') as file:
				file = file.readlines()
				for element in file:
					report.write(element)
			report.write("\n")
		"""if os.path.exists("../Results/"+REPERTORYVCF+"/temp/Polymorphism_"+i+".vcf")== True:
			report.write("Polymorphism:")
			report.write("\n")
			File = "../Results/"+REPERTORYVCF+"/temp/Polymorphism_"+i+".vcf"
			with open(File,'r') as file:
				file = file.readlines()
				for element in file:
					report.write(element)
		report.write("\n")"""
		if os.path.exists(RESULTDIR+"/"+REPERTORYVCF+"/temp/no_contributory_"+i+".vcf")== True:
			report.write("Variants détectés mais non retenus: (NOCALL ou allele_freq < 1%  ou < 25 reads)")
			report.write("\n")
			File = RESULTDIR+"/"+REPERTORYVCF+"/temp/no_contributory_"+i+".vcf"
			with open(File,'r') as file:
				file = file.readlines()
				for element in file:
					report.write(element)
			report.write("\n")

		#rajouter definitions
		report.write("Définitions:\n")
		report.write("SIFT : Prediction de l'impact d'une substitution sur la fonction de la protéine en se basant sur le degré de conservation des acides aminés\n")
		report.write("Polyphen : Prédiction de l'impact d'une substitution sur la structure et la fonction des protéines en utilisant des informations de séquences et de structure\n")
		report.write("HGVSc: le nom de la séquence codante dans HGVS (Human Genome Variation Society)\n")
		report.write("HGVSp: le nom de la séquence protéique dans HGVS (Human Genome Variation Society)\n")
		report.write("Profondeur: La profondeur de lecture est le nombre de lectures («reads») indépendantes d'une base par le séquenceur\n")
		report.write("Couverture: La couverture (en pourcentage) est le nombre de bases couvertes par rapport au nombre total de bases de la région d'intérêt (pour une profondeur de lecture donnée)\n")
		report.write("maf: fréquence de l'allèle minoritaire\n")


		report.close()

	def informations(self,REPERTORYVCF,i,RESULTDIR,pathREPERTORYVCF):
		if os.path.exists(pathREPERTORYVCF+"/templateNGS.txt") == True:
			info=""
			number = i[10:]
			if "0" in number:
				number = number.replace("0","")
			File = pathREPERTORYVCF+"/templateNGS.txt"
			#with open(File,'r') as file:
			with codecs.open(File, "r",encoding='utf-8', errors='ignore') as file:
				file = file.readlines()
				for element in file:
					fileSplit = element.split(",")
					if fileSplit[0] == number :
						info = "Identifiant = "+fileSplit[1]+"\nNom = "+fileSplit[2]+"\nIndication = "+fileSplit[3]+"\nPanel = "+fileSplit[4]+"\n" 
			return info


	def pyxl(self,i, REPERTORYVCF,RESULTDIR):
		wb = Workbook()
		ws = wb.active
		ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
		ws.page_setup.paperSize = ws.PAPERSIZE_TABLOID
		#####
		# Largeur des colonnes
		#####
		#Pour le sinformations:
		ws.column_dimensions["A"].width = 10.0
		ws.column_dimensions["B"].width =12.0		
		ws.column_dimensions["C"].width = 14.0
		ws.column_dimensions["D"].width =10.0		
		ws.column_dimensions["E"].width =12.0		
		ws.column_dimensions["F"].width =12.0
		ws.column_dimensions["G"].width = 10.0
		ws.column_dimensions["H"].width = 12.0
		ws.column_dimensions["I"].width = 12.0
		ws.column_dimensions["J"].width = 12.0
		ws.column_dimensions["K"].width = 10.0
		ws.column_dimensions["L"].width =7.0
		ws.column_dimensions["M"].width = 7.0
		ws.column_dimensions["N"].width = 7.0
				#Titre du fichier
		if i[-3:]=="vcf":
			i=i[:-4]
		titre=i
		ws.title = titre

		fichier = open(RESULTDIR+"/"+REPERTORYVCF+"/temp/Report_info_"+i+".txt","r")
		content = fichier.readlines()
		border_thin = Border(left=Side(style='thin',color='FF000000'),right=Side(style='thin',color='FF000000'),top=Side(style='thin',color='FF000000'),bottom=Side(style='thin',color='FF000000'))
		alpahabet = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W"]

		for row in range(len(content)):
			contentrowsplit = content[row].split("\t")
			contentrowsplit1 = content[row].split("\t")
			# Met en gras les legendes
			Gras=False
			if contentrowsplit[0] == "Gene" or contentrowsplit[0] == "Echantillon":
				Gras=True
				
			#else:
				#font = Font(name='Arial',size=8, bold=False)
			for col in range(len(contentrowsplit)):
				if len(contentrowsplit) == 1:
					if "=" in contentrowsplit[0]:
						font = Font(name='Arial',size=12, bold=False)
						cell = "A"+str(row+1)
						ws[cell] = content[row]
						ws[cell].font = font
					else:
						font = Font(name='Arial',size=12, bold=True, underline='single')
						cell = "A"+str(row+1)
						ws[cell] = content[row]
						ws[cell].font = font

				else:
					if Gras:
						font = Font(name='Arial',size=10, bold=True)

					else:
						font = Font(name='Arial',size=10, bold=False)
					if ws.column_dimensions[alpahabet[col]].width != None:
						listeHeight = []
						if ws.column_dimensions[alpahabet[col]].width < len(contentrowsplit[col]):
								if len(contentrowsplit[col])>15:

									if Gras:
										font = Font(name='Arial',size=10, bold=True)
										contentrowsplit[col]=str(contentrowsplit[col][:7])+str(contentrowsplit[col][7:]).replace(" ","\n",1)
										ws.row_dimensions[row+1].height = 30
									else:
										font = Font(name='Arial',size=10, bold=False)
										a=1
										height=20
										contentrowsplit1[col]=contentrowsplit[col]
										for caractere in contentrowsplit[col]:
											a+=1
											if a%12==True:
												contentrowsplit1[col]=contentrowsplit1[col][:a]+"\n"+contentrowsplit1[col][a:]
												height+=20
										listeHeight.append(height)
										ws.row_dimensions[row+1].height = max(listeHeight)
										contentrowsplit[col]=contentrowsplit1[col]
						
						if col == len(contentrowsplit)-1:
							cell = alpahabet[col]+str(row+1)
							#print(cell)
							ws[cell] = contentrowsplit[col]
							ws[cell].font = font
							ws[cell].alignment = Alignment(horizontal="center",vertical="center")
						else:
							cell = alpahabet[col]+str(row+1)
							ws[cell] = contentrowsplit[col]
							ws[cell].font = font
							ws[cell].border = border_thin
							ws[cell].alignment = Alignment(horizontal="center",vertical="center")


		#Pour les variants:
		ws1 = wb.create_sheet(title="Variants")
		ws1.page_setup.orientation = ws1.ORIENTATION_LANDSCAPE
		ws1.page_setup.paperSize = ws1.PAPERSIZE_TABLOID
		#####
		# Largeur des colonnes
		#####
		#Pour le sinformations:
		ws1.column_dimensions["A"].width = 8.0
		ws1.column_dimensions["B"].width =0# 10.0		
		ws1.column_dimensions["C"].width = 10.0
		ws1.column_dimensions["D"].width =10		
		ws1.column_dimensions["E"].width =15# 12.0		
		ws1.column_dimensions["F"].width =13# 10.0
		ws1.column_dimensions["G"].width = 10
		ws1.column_dimensions["H"].width = 8.0
		ws1.column_dimensions["I"].width = 10
		ws1.column_dimensions["J"].width = 0
		ws1.column_dimensions["K"].width = 12
		ws1.column_dimensions["L"].width =10
		ws1.column_dimensions["M"].width =0 
		ws1.column_dimensions["N"].width =9.0


		

		fichier2 = open(RESULTDIR+"/"+REPERTORYVCF+"/temp/Report_"+i+".txt","r")
		content = fichier2.readlines()
		border_thin = Border(left=Side(style='thin',color='FF000000'),right=Side(style='thin',color='FF000000'),top=Side(style='thin',color='FF000000'),bottom=Side(style='thin',color='FF000000'))
		alpahabet = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF","AG","AH","AI","AJ","AK","AL","AM","AN","AO","AP","AQ","AR","AS","AT","AU","AV","AW","AX","AY","AZ","BA","BB","BC","BD","BE","BF","BG","BH","BI","BJ","BK","BL","BM","BN","BO","BP","BQ","BR","BS","BT","BU","BV","BW","BX","BY","BZ"]
		NotDef=True
		for row in range(len(content)):
			contentrowsplit = content[row].split("\t")
			contentrowsplit1 = content[row].split("\t")	
			# Met en gras les legendes
			Gras=False
			if contentrowsplit[0] == "Gene" or contentrowsplit[0] == "Echantillon":
				Gras=True
			listeHeight = []
			for col in range(len(contentrowsplit)):
				if len(contentrowsplit) == 1:
					if NotDef:
						font = Font(name='Arial',size=10, bold=True, underline='single')
						cell = "A"+str(row+1)
						ws1[cell] = content[row]
						ws1[cell].font = font
						if str(content[row][:12])=="Définitions:":
							NotDef=False

					else:
						font = Font(name='Arial',size=8, bold=False)
						cell = "A"+str(row+1)
						ws1[cell] = content[row]
						ws1[cell].font = font

				else:
					if Gras:
						font = Font(name='Arial',size=10, bold=True)
					else:
						font = Font(name='Arial',size=10, bold=False)
					if ws1.column_dimensions[alpahabet[col]].width != None:						
						if ws1.column_dimensions[alpahabet[col]].width < len(contentrowsplit[col]):
							
							if len(contentrowsplit[col])>15:
								
								if Gras:
									font = Font(name='Arial',size=10, bold=True)
									contentrowsplit[col]=str(contentrowsplit[col][:7])+str(contentrowsplit[col][7:]).replace(" ","\n",1)
									ws1.row_dimensions[row+1].height = 20
								else:
									font = Font(name='Arial',size=10, bold=False)
									a=1
									height=20
									contentrowsplit1[col]=contentrowsplit[col]
									for caractere in contentrowsplit[col]:
										a+=1
										if a%12==True:
											contentrowsplit1[col]=contentrowsplit1[col][:a]+"\n"+contentrowsplit1[col][a:]
											height+=20
									listeHeight.append(height)
									ws1.row_dimensions[row+1].height = max(listeHeight)
									contentrowsplit[col]=contentrowsplit1[col]
						if col == len(contentrowsplit)-1:
							cell = alpahabet[col]+str(row+1)
							#print(cell)
							ws1[cell] = contentrowsplit[col]
							ws1[cell].font = font
							ws1[cell].alignment = Alignment(horizontal="center",vertical="center")
						else:
							cell = alpahabet[col]+str(row+1)
							ws1[cell] = contentrowsplit[col]
							ws1[cell].font = font
							ws1[cell].border = border_thin
							ws1[cell].alignment = Alignment(horizontal="center",vertical="center")
		
		wb.save(RESULTDIR+"/"+REPERTORYVCF+"/Report_"+i+".xlsx")
		a = "libreoffice --headless --invisible --convert-to pdf --outdir "+RESULTDIR+"/"+REPERTORYVCF+" "+RESULTDIR+"/"+REPERTORYVCF+"/Report_"+i+".xlsx"
		os.system(a)
		mycwd = os.getcwd()
		os.chdir(RESULTDIR+"/"+REPERTORYVCF+"/")
		c = "tar cf Report_"+i+".tar Report_"+i+".xlsx Report_"+i+".pdf"
		os.system(c)
		os.remove("Report_"+i+".xlsx")
		os.remove("Report_"+i+".pdf")
		os.chdir(mycwd)




